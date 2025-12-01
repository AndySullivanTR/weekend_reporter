from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from datetime import datetime, timedelta
import json
import os
from werkzeug.security import generate_password_hash, check_password_hash
import secrets
import random
try:
    import fcntl
    HAS_FCNTL = True
except ImportError:
    # fcntl is not available on Windows
    HAS_FCNTL = False

# Determine the base directory (where this script is located)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(BASE_DIR)

template_folder = os.path.join(BASE_DIR, 'templates')
app = Flask(__name__, template_folder=template_folder)

# Fixed secret key for session persistence across restarts
app.secret_key = 'weekend-reporter-shifts-secret-key-2025'

# Data storage (in production, use a proper database)
DATA_DIR = os.path.join(BASE_DIR, 'data')
os.makedirs(DATA_DIR, exist_ok=True)

BACKUP_DIR = os.path.join(DATA_DIR, 'backups')
os.makedirs(BACKUP_DIR, exist_ok=True)

REPORTERS_FILE = os.path.join(DATA_DIR, 'reporters.json')
PREFERENCES_FILE = os.path.join(DATA_DIR, 'preferences.json')
SETTINGS_FILE = os.path.join(DATA_DIR, 'settings.json')
ASSIGNMENTS_FILE = os.path.join(DATA_DIR, 'assignments.json')

# Generate 84 weekend shifts (21 weekends starting Dec 13, 2025)
# 4 shifts per weekend: Sat morning, Sat evening, Sun morning, Sun evening
# Total capacity: 126 slots (for 123 reporters - excludes TEST account)
def generate_shifts():
    shifts = []
    start_date = datetime(2025, 12, 13)  # Saturday Dec 13, 2025
    shift_id = 0
    
    for week in range(21):
        saturday = start_date + timedelta(weeks=week)
        sunday = saturday + timedelta(days=1)
        
        # Saturday morning shift - 1 reporter
        shifts.append({
            'id': shift_id,
            'date': saturday.strftime('%Y-%m-%d'),
            'day': 'Saturday',
            'time': '8:00 AM - 4:00 PM',
            'slots': 1,
            'week': week + 1
        })
        shift_id += 1
        
        # Saturday evening shift - 1 reporter
        shifts.append({
            'id': shift_id,
            'date': saturday.strftime('%Y-%m-%d'),
            'day': 'Saturday',
            'time': '3:00 PM - 10:00 PM',
            'slots': 1,
            'week': week + 1
        })
        shift_id += 1
        
        # Sunday morning shift - 2 reporters
        shifts.append({
            'id': shift_id,
            'date': sunday.strftime('%Y-%m-%d'),
            'day': 'Sunday',
            'time': '8:00 AM - 4:00 PM',
            'slots': 2,
            'week': week + 1
        })
        shift_id += 1
        
        # Sunday evening shift - 2 reporters
        shifts.append({
            'id': shift_id,
            'date': sunday.strftime('%Y-%m-%d'),
            'day': 'Sunday',
            'time': '3:00 PM - 10:00 PM',
            'slots': 2,
            'week': week + 1
        })
        shift_id += 1
    
    return shifts

SHIFTS = generate_shifts()

# Initialize data files
def init_data_files():
    # Create 123 reporters (use reload-reporters-from-csv endpoint to load actual credentials)
    if not os.path.exists(REPORTERS_FILE):
        reporters = {}
        
        # Manager account
        reporters['admin'] = {
            'name': 'Admin',
            'is_manager': True,
            'password': generate_password_hash('admin123')
        }
        
        # 123 reporter accounts (placeholder - use reload endpoint for real credentials)
        for i in range(1, 124):
            username = f'reporter{i}'
            reporters[username] = {
                'name': f'Reporter{i}',
                'is_manager': False,
                'password': generate_password_hash('password')
            }
        
        with open(REPORTERS_FILE, 'w') as f:
            json.dump(reporters, f, indent=2)
    
    if not os.path.exists(PREFERENCES_FILE):
        with open(PREFERENCES_FILE, 'w') as f:
            json.dump({}, f)
    
    if not os.path.exists(SETTINGS_FILE):
        # Default deadline: 7 days from now
        deadline = (datetime.now() + timedelta(days=7)).isoformat()
        with open(SETTINGS_FILE, 'w') as f:
            json.dump({'deadline': deadline, 'is_locked': False}, f)
    
    if not os.path.exists(ASSIGNMENTS_FILE):
        with open(ASSIGNMENTS_FILE, 'w') as f:
            json.dump({}, f)

init_data_files()

# Helper functions
def load_json(filepath):
    with open(filepath, 'r') as f:
        return json.load(f)

def save_json(filepath, data):
    """Save JSON with file locking to prevent race conditions (Unix only)"""
    with open(filepath, 'w') as f:
        if HAS_FCNTL:
            # Acquire exclusive lock (Unix only)
            fcntl.flock(f.fileno(), fcntl.LOCK_EX)
        try:
            json.dump(data, f, indent=2)
        finally:
            if HAS_FCNTL:
                # Release lock
                fcntl.flock(f.fileno(), fcntl.LOCK_UN)

def get_reporters():
    return load_json(REPORTERS_FILE)

def get_preferences():
    return load_json(PREFERENCES_FILE)

def get_settings():
    return load_json(SETTINGS_FILE)

def get_assignments():
    return load_json(ASSIGNMENTS_FILE)

def create_auto_backup():
    """Create an automatic backup of all data files"""
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_data = {
            'reporters': get_reporters(),
            'preferences': get_preferences(),
            'settings': get_settings(),
            'assignments': get_assignments(),
            'timestamp': datetime.now().isoformat()
        }
        
        backup_file = os.path.join(BACKUP_DIR, f'auto_backup_{timestamp}.json')
        with open(backup_file, 'w') as f:
            json.dump(backup_data, f, indent=2)
        
        # Keep only last 30 backups to save space
        backup_files = sorted([f for f in os.listdir(BACKUP_DIR) if f.startswith('auto_backup_')])
        if len(backup_files) > 30:
            for old_backup in backup_files[:-30]:
                os.remove(os.path.join(BACKUP_DIR, old_backup))
        
        return True
    except Exception as e:
        print(f"Auto-backup failed: {e}")
        return False

def format_deadline(iso_datetime_str):
    """Format ISO datetime to readable format: 'Nov. 27, 2025 2:37 pm ET'"""
    dt = datetime.fromisoformat(iso_datetime_str.replace('Z', '+00:00'))
    if dt.tzinfo is not None:
        dt = dt.replace(tzinfo=None)
    
    month = dt.strftime('%b')
    day = dt.day
    year = dt.year
    hour = dt.hour
    minute = dt.minute
    
    # Convert to 12-hour format with am/pm
    if hour == 0:
        hour_12 = 12
        am_pm = 'am'
    elif hour < 12:
        hour_12 = hour
        am_pm = 'am'
    elif hour == 12:
        hour_12 = 12
        am_pm = 'pm'
    else:
        hour_12 = hour - 12
        am_pm = 'pm'
    
    return f"{month}. {day}, {year} {hour_12}:{minute:02d} {am_pm} ET"

# Routes
@app.route('/')
def index():
    if 'username' in session:
        if session.get('is_manager'):
            return redirect(url_for('manager_dashboard'))
        else:
            return redirect(url_for('reporter_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.json
        username = data.get('username')
        password = data.get('password')
        
        reporters = get_reporters()
        
        if username in reporters:
            if check_password_hash(reporters[username]['password'], password):
                session['username'] = username
                session['is_manager'] = reporters[username].get('is_manager', False)
                return jsonify({'success': True, 'is_manager': session['is_manager']})
        
        return jsonify({'success': False, 'message': 'Invalid credentials'}), 401
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/manager/dashboard')
def manager_dashboard():
    if not session.get('is_manager'):
        return redirect(url_for('login'))
    
    reporters = get_reporters()
    settings = get_settings()
    preferences = get_preferences()
    assignments = get_assignments()
    
    # Count reporters who have submitted preferences (top 10 + bottom 5)
    submitted_count = sum(1 for rep, prefs in preferences.items() 
                         if prefs and len(prefs.get('top_10', [])) == 10 and len(prefs.get('bottom_5', [])) == 5)
    
    return render_template('manager_dashboard.html', 
                         reporters=reporters,
                         settings=settings,
                         submitted_count=submitted_count,
                         total_reporters=len([r for r in reporters.values() if not r.get('is_manager')]),
                         assignments=assignments,
                         preferences=preferences,
                         shifts=SHIFTS)

@app.route('/reporter/dashboard')
def reporter_dashboard():
    if 'username' not in session or session.get('is_manager'):
        return redirect(url_for('login'))
    
    settings = get_settings()
    preferences = get_preferences()
    assignments = get_assignments()
    username = session['username']
    
    user_prefs = preferences.get(username, {})
    user_assignments = assignments.get(username, [])
    
    # Check if locked (no deadline check)
    is_locked = settings.get('is_locked', False)
    formatted_deadline = 'No deadline'
    
    return render_template('reporter_dashboard.html',
                         username=username,
                         shifts=SHIFTS,
                         preferences=user_prefs,
                         assignments=user_assignments,
                         deadline=formatted_deadline,
                         is_locked=is_locked)

@app.route('/api/preferences', methods=['GET', 'POST'])
def manage_preferences():
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 403
    
    username = session['username']
    preferences = get_preferences()
    settings = get_settings()
    
    # Check if locked (no deadline check)
    is_locked = settings.get('is_locked', False)
    
    if request.method == 'POST':
        if is_locked and not session.get('is_manager'):
            return jsonify({'error': 'Preferences are locked'}), 403
        
        data = request.json
        
        # Validate data structure
        if 'top_10' not in data or 'bottom_5' not in data or 'shift_type_pref' not in data:
            return jsonify({'error': 'Invalid preference format'}), 400
        
        if len(data['top_10']) != 10:
            return jsonify({'error': 'Must select exactly 10 top preferences'}), 400
        
        if len(data['bottom_5']) != 5:
            return jsonify({'error': 'Must select exactly 5 least wanted shifts'}), 400
        
        preferences[username] = {
            'top_10': data['top_10'],
            'bottom_5': data['bottom_5'],
            'shift_type_pref': data['shift_type_pref']
        }
        save_json(PREFERENCES_FILE, preferences)
        
        # Create auto-backup after preference submission
        create_auto_backup()
        
        return jsonify({'success': True})
    
    # GET
    if session.get('is_manager'):
        return jsonify(preferences)
    else:
        return jsonify({username: preferences.get(username, {})})

@app.route('/api/settings', methods=['GET', 'POST'])
def manage_settings():
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    settings = get_settings()
    
    if request.method == 'POST':
        data = request.json
        
        if 'deadline' in data:
            settings['deadline'] = data['deadline']
        
        if 'is_locked' in data:
            settings['is_locked'] = data['is_locked']
        
        save_json(SETTINGS_FILE, settings)
        return jsonify({'success': True})
    
    return jsonify(settings)

@app.route('/api/allocate', methods=['POST'])
def allocate_shifts():
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    # Create backup before allocation
    create_auto_backup()
    
    preferences = get_preferences()
    reporters_data = get_reporters()
    
    # Get list of non-manager reporters
    reporter_list = [user for user, rep in reporters_data.items() if not rep.get('is_manager') and user != 'test']
    
    # Separate reporters into those WITH and WITHOUT preferences
    reporters_with_prefs = []
    reporters_without_prefs = []
    warnings = []
    
    for rep in reporter_list:
        if rep in preferences:
            prefs = preferences[rep]
            if len(prefs.get('top_10', [])) == 10 and len(prefs.get('bottom_5', [])) == 5:
                reporters_with_prefs.append(rep)
            else:
                reporters_without_prefs.append(rep)
                rep_name = reporters_data[rep]['name']
                warnings.append(f"{rep_name} has incomplete preferences - will be randomly assigned")
        else:
            reporters_without_prefs.append(rep)
            rep_name = reporters_data[rep]['name']
            warnings.append(f"{rep_name} did not submit preferences - will be randomly assigned")
    
    # Initialize assignments
    assignments = {rep: [] for rep in reporter_list}
    shift_assignments = {shift['id']: [] for shift in SHIFTS}
    
    # Set random seed for reproducibility
    random.seed(42)
    
    # PHASE 1: Allocate for reporters WITH preferences
    # Strategy: Fill weeks 1-20 first (shifts 0-59), then week 21 (shifts 60-62)
    print("\n=== REPORTER SHIFT ALLOCATION (WITH PREFERENCES) ===")
    print("Strategy: Prioritizing weeks 1-20 (shifts 0-59) to be fully filled")
    
    shuffled_reporters = reporters_with_prefs.copy()
    random.shuffle(shuffled_reporters)
    
    # Assign one shift to each reporter
    for rep in shuffled_reporters:
        prefs = preferences[rep]
        top_10 = prefs['top_10']
        bottom_5 = prefs['bottom_5']
        
        # Try to assign from top 10 preferences, prioritizing weeks 1-20
        assigned = False
        
        # First try: top 10 preferences that are in weeks 1-20 (shift IDs 0-59)
        for shift_id in top_10:
            if shift_id >= 60:  # Skip week 21 for now
                continue
            
            shift = next(s for s in SHIFTS if s['id'] == shift_id)
            if len(shift_assignments[shift_id]) >= shift['slots']:
                continue
            
            assignments[rep].append(shift_id)
            shift_assignments[shift_id].append(rep)
            assigned = True
            rank = top_10.index(shift_id) + 1
            print(f"✓ {rep:30} → Shift {shift_id:2} (week {shift['week']}, preference #{rank})")
            break
        
        # Second try: top 10 preferences in week 21 if nothing in weeks 1-20
        if not assigned:
            for shift_id in top_10:
                if shift_id < 60:  # Already tried these
                    continue
                
                shift = next(s for s in SHIFTS if s['id'] == shift_id)
                if len(shift_assignments[shift_id]) >= shift['slots']:
                    continue
                
                assignments[rep].append(shift_id)
                shift_assignments[shift_id].append(rep)
                assigned = True
                rank = top_10.index(shift_id) + 1
                print(f"✓ {rep:30} → Shift {shift_id:2} (week 21, preference #{rank})")
                break
        
        # Third try: non-bottom-5 shifts in weeks 1-20
        if not assigned:
            shift_type_pref = prefs.get('shift_type_pref', {})
            sorted_types = sorted(shift_type_pref.items(), key=lambda x: x[1])
            
            for shift_type, _ in sorted_types:
                for shift in SHIFTS:
                    shift_id = shift['id']
                    
                    # Prioritize weeks 1-20
                    if shift_id >= 60:
                        continue
                    
                    if shift_id in bottom_5 or shift_id in top_10:
                        continue
                    
                    # Check shift type match
                    shift_matches = False
                    if shift_type == 'saturday' and shift['day'] == 'Saturday':
                        shift_matches = True
                    elif shift_type == 'sunday_morning' and shift['day'] == 'Sunday' and '8:00 AM' in shift['time']:
                        shift_matches = True
                    elif shift_type == 'sunday_evening' and shift['day'] == 'Sunday' and '3:00 PM' in shift['time']:
                        shift_matches = True
                    
                    if not shift_matches:
                        continue
                    
                    if len(shift_assignments[shift_id]) >= shift['slots']:
                        continue
                    
                    assignments[rep].append(shift_id)
                    shift_assignments[shift_id].append(rep)
                    assigned = True
                    print(f"⚠ {rep:30} → Shift {shift_id:2} (week {shift['week']}, backup assignment)")
                    break
                
                if assigned:
                    break
        
        if not assigned:
            print(f"✗ {rep:30} → Could not assign shift")
    
    # PHASE 2: Random allocation for reporters WITHOUT preferences
    if reporters_without_prefs:
        print("\n=== RANDOM ALLOCATION (NO PREFERENCES) ===")
        
        # Assign 1 shift to each reporter, prioritizing weeks 1-20
        for rep in reporters_without_prefs:
            # Create pool of available shifts, prioritizing weeks 1-20 (shifts 0-59)
            available_weeks_1_20 = []
            available_week_21 = []
            
            for shift in SHIFTS:
                shift_id = shift['id']
                filled = len(shift_assignments[shift_id])
                capacity = shift['slots']
                
                if filled < capacity:
                    if shift_id < 60:  # Weeks 1-20
                        available_weeks_1_20.append(shift_id)
                    else:  # Week 21
                        available_week_21.append(shift_id)
            
            # Try weeks 1-20 first
            if available_weeks_1_20:
                random.shuffle(available_weeks_1_20)
                shift_id = available_weeks_1_20[0]
                shift = next(s for s in SHIFTS if s['id'] == shift_id)
                assignments[rep].append(shift_id)
                shift_assignments[shift_id].append(rep)
                print(f"🎲 {rep:30} → Shift {shift_id:2} (week {shift['week']}, random)")
            elif available_week_21:
                # Only use week 21 if weeks 1-20 are full
                random.shuffle(available_week_21)
                shift_id = available_week_21[0]
                shift = next(s for s in SHIFTS if s['id'] == shift_id)
                assignments[rep].append(shift_id)
                shift_assignments[shift_id].append(rep)
                print(f"🎲 {rep:30} → Shift {shift_id:2} (week 21, random)")
            else:
                print(f"✗ {rep:30} → No shifts available")
                warnings.append(f"{reporters_data[rep]['name']} could not be assigned - no capacity remaining")
    
    # Save assignments
    save_json(ASSIGNMENTS_FILE, assignments)
    
    # Lock preferences
    settings = get_settings()
    settings['is_locked'] = True
    save_json(SETTINGS_FILE, settings)
    
    return jsonify({
        'success': True,
        'assignments': assignments,
        'shift_assignments': shift_assignments,
        'warnings': warnings,
        'reporters_with_prefs': len(reporters_with_prefs),
        'reporters_without_prefs': len(reporters_without_prefs)
    })

@app.route('/api/backup')
def backup_data():
    """Download all data files as JSON for backup"""
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    backup_data = {
        'reporters': get_reporters(),
        'preferences': get_preferences(),
        'settings': get_settings(),
        'assignments': get_assignments(),
        'timestamp': datetime.now().isoformat()
    }
    
    from io import BytesIO
    output = BytesIO()
    output.write(json.dumps(backup_data, indent=2).encode('utf-8'))
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/json',
        as_attachment=True,
        download_name=f'backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
    )

@app.route('/api/populate-test-data', methods=['POST'])
def populate_test_data():
    """Populate random preferences for all reporters (TESTING ONLY)"""
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    reporters = get_reporters()
    preferences = {}
    
    # All 60 shift IDs (weeks 1-20, not week 21)
    all_shifts = list(range(60))
    
    # Generate random preferences for each non-manager reporter
    for username, rep_data in reporters.items():
        if rep_data.get('is_manager') or username == 'test':
            continue
        
        # Shuffle all shifts
        shuffled = all_shifts.copy()
        random.shuffle(shuffled)
        
        # Top 10 are first 10 from shuffled list
        top_10 = shuffled[:10]
        
        # Bottom 5 are next 5 from shuffled list
        bottom_5 = shuffled[10:15]
        
        # Random shift type preferences (1, 2, 3)
        shift_types = [1, 2, 3]
        random.shuffle(shift_types)
        
        preferences[username] = {
            'top_10': top_10,
            'bottom_5': bottom_5,
            'shift_type_pref': {
                'saturday': str(shift_types[0]),
                'sunday_morning': str(shift_types[1]),
                'sunday_evening': str(shift_types[2])
            }
        }
    
    # Save preferences
    save_json(PREFERENCES_FILE, preferences)
    
    return jsonify({
        'success': True,
        'message': f'Populated random preferences for {len(preferences)} reporters'
    })

@app.route('/api/export-excel')
def export_excel():
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        assignments = get_assignments()
        reporters = get_reporters()
        preferences = get_preferences()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporter Schedule"
        
        # Title
        ws['A1'] = 'Weekend Reporter Shift Schedule - Dec 2025 - Apr 2026'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')
        
        # Headers
        headers = ['Date', 'Day', 'Time', 'Assigned Reporters', 'Preference Rank', 'Status', 'Week', 'Notes']
        header_row = 3
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        row = header_row + 1
        for shift in SHIFTS:
            shift_id = shift['id']
            assigned = []
            
            for rep, rep_shifts in assignments.items():
                if shift_id in rep_shifts:
                    assigned.append(rep)
            
            ws.cell(row=row, column=1).value = shift['date']
            ws.cell(row=row, column=2).value = shift['day']
            ws.cell(row=row, column=3).value = shift['time']
            
            # Assigned reporters (can be 0, 1, or 2)
            if assigned:
                rep_names = []
                pref_ranks = []
                for rep in assigned:
                    rep_names.append(reporters[rep]['name'])
                    
                    if rep in preferences:
                        prefs = preferences[rep]
                        if shift_id in prefs.get('top_10', []):
                            rank = prefs['top_10'].index(shift_id) + 1
                            pref_ranks.append(f"#{rank}")
                        elif shift_id in prefs.get('bottom_5', []):
                            pref_ranks.append("Bottom-5")
                        else:
                            pref_ranks.append("N/A")
                    else:
                        pref_ranks.append("N/A")
                
                ws.cell(row=row, column=4).value = ", ".join(rep_names)
                ws.cell(row=row, column=5).value = ", ".join(pref_ranks)
            else:
                ws.cell(row=row, column=4).value = "VACANT"
                ws.cell(row=row, column=4).font = Font(color="FF0000", bold=True)
            
            # Status
            filled = len(assigned)
            total = shift['slots']
            if filled >= total:
                ws.cell(row=row, column=6).value = "FILLED"
                ws.cell(row=row, column=6).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                ws.cell(row=row, column=6).value = f"VACANT ({total - filled})"
                ws.cell(row=row, column=6).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            ws.cell(row=row, column=7).value = shift['week']
            
            row += 1
        
        # Reporter summary
        row += 2
        ws.cell(row=row, column=1).value = "Reporter Summary"
        ws.cell(row=row, column=1).font = Font(size=14, bold=True)
        
        row += 1
        summary_headers = ['Reporter', 'Shifts Assigned', 'Shift Details', 'Status']
        for col, header in enumerate(summary_headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        
        row += 1
        for rep, rep_data in reporters.items():
            if rep_data.get('is_manager'):
                continue
            
            ws.cell(row=row, column=1).value = rep_data['name']
            
            rep_shifts = assignments.get(rep, [])
            ws.cell(row=row, column=2).value = len(rep_shifts)
            
            shift_details = []
            for shift_id in rep_shifts:
                shift = next(s for s in SHIFTS if s['id'] == shift_id)
                shift_details.append(f"{shift['date']} {shift['day']} {shift['time']}")
            ws.cell(row=row, column=3).value = "; ".join(shift_details) if shift_details else "None"
            
            if len(rep_shifts) == 1:
                ws.cell(row=row, column=4).value = "Complete"
                ws.cell(row=row, column=4).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                ws.cell(row=row, column=4).value = f"Incomplete ({len(rep_shifts)}/1)"
                ws.cell(row=row, column=4).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 30
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'reporter_schedule_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/change-password', methods=['POST'])
def change_password():
    """Allow users to change their password"""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 403
    
    username = session['username']
    data = request.json
    
    current_password = data.get('current_password')
    new_password = data.get('new_password')
    
    if not current_password or not new_password:
        return jsonify({'error': 'Current and new password required'}), 400
    
    if len(new_password) < 6:
        return jsonify({'error': 'New password must be at least 6 characters'}), 400
    
    reporters = get_reporters()
    
    # Verify current password
    if username not in reporters:
        return jsonify({'error': 'User not found'}), 404
    
    if not check_password_hash(reporters[username]['password'], current_password):
        return jsonify({'error': 'Current password is incorrect'}), 401
    
    # Update password
    reporters[username]['password'] = generate_password_hash(new_password)
    save_json(REPORTERS_FILE, reporters)
    
    return jsonify({'success': True, 'message': 'Password changed successfully'})

# Reporter credentials data (embedded to avoid CSV file dependency)
REPORTER_CREDENTIALS = [
    {"name": "Aboulenein, Ahmed", "username": "ahmed.aboulenein", "password": "isrcoi"},
    {"name": "Ahmed, Saqib", "username": "saqib.ahmed", "password": "VtaIrc"},
    {"name": "Alleyne-Morris, Shawana", "username": "shawana.alleyne-morris", "password": "505eE3"},
    {"name": "Anand, Nupur", "username": "nupur.anand", "password": "OGbF9J"},
    {"name": "Azhar, Saeed", "username": "saeed.azhar", "password": "m7Is0j"},
    {"name": "Baertlein, Lisa P.", "username": "lisa.baertlein", "password": "3Ju3Wb"},
    {"name": "Banco, Erin", "username": "erin.banco", "password": "wMkhw0"},
    {"name": "Barbuscia, Davide", "username": "davide.barbuscia", "password": "SOumy4"},
    {"name": "Bautzer, Tatiana", "username": "tatiana.bautzer", "password": "x8ZQRd"},
    {"name": "Bensinger, Greg", "username": "greg.bensinger", "password": "OxzLbT"},
    {"name": "Binnie, Isla", "username": "isla.binnie", "password": "MxO3sq"},
    {"name": "Brettell, Karen J.", "username": "karen.brettell", "password": "5Wfwcu"},
    {"name": "Brittain, Blake", "username": "blake.brittain", "password": "83RomE"},
    {"name": "Brown, Nicholas P.", "username": "nicholas.p.brown", "password": "eJTLFQ"},
    {"name": "Cai, Kenrick", "username": "kenrick.cai", "password": "H66q4i"},
    {"name": "Campos, Rodrigo", "username": "rodrigo.campos", "password": "gXTzhk"},
    {"name": "Carew, Sinead M.", "username": "sinead.carew", "password": "QGAiY5"},
    {"name": "Catchpole, Dan", "username": "dan.catchpole", "password": "Cgcp7n"},
    {"name": "Cavale, Siddharth", "username": "siddharth.cavale", "password": "ncz2NL"},
    {"name": "Chavez, Gertrude", "username": "gertrude.chavez", "password": "NOWLwL"},
    {"name": "Cherney, Max A.", "username": "max.cherney", "password": "HEP1Ay"},
    {"name": "Chmielewski, Dawn C.", "username": "dawn.chmielewski", "password": "sFjwWj"},
    {"name": "Chung, Andrew", "username": "andrew.chung", "password": "sulpg2"},
    {"name": "Cohen, Luc", "username": "luc.cohen", "password": "tLZt8t"},
    {"name": "Conlin, Michelle", "username": "michelle.conlin", "password": "Ne29uB"},
    {"name": "Cooke, Kristina R.", "username": "kristina.cooke", "password": "O2nUzm"},
    {"name": "Culp, Stephen R.", "username": "stephen.culp", "password": "vyOW3V"},
    {"name": "Cunningham, Waylon", "username": "waylon.cunningham", "password": "5eS7RJ"},
    {"name": "Dang, Sheila", "username": "sheila.dang", "password": "AmxX4u"},
    {"name": "Dastin, Jeffrey", "username": "jeffrey.dastin", "password": "mAnP4c"},
    {"name": "Delevingne, Lawrence", "username": "lawrence.delevingne", "password": "QAoEDS"},
    {"name": "Derby, Michael", "username": "michael.derby", "password": "AV2MnB"},
    {"name": "DiNapoli, Jessica", "username": "jessica.dinapoli", "password": "N60hyD"},
    {"name": "DiSavino, Scott P.", "username": "scott.disavino", "password": "XAIeq9"},
    {"name": "Douglas, Leah", "username": "leah.douglas", "password": "023vja"},
    {"name": "Eckert, Nora", "username": "nora.eckert", "password": "nPIwnl"},
    {"name": "Erman, Michael D.", "username": "michael.erman", "password": "gPrCnL"},
    {"name": "Flowers, Bianca", "username": "bianca.flowers", "password": "fxq9qp"},
    {"name": "Freifeld, Karen", "username": "karen.freifeld", "password": "dLo9IV"},
    {"name": "French, David J.", "username": "davidj.french", "password": "nI9MDy"},
    {"name": "Gardner, Timothy", "username": "timothy.gardner", "password": "U70mDI"},
    {"name": "Gillison, Douglas", "username": "douglas.gillison", "password": "2idb2J"},
    {"name": "Godoy, Jody", "username": "jody.godoy", "password": "UrU0eP"},
    {"name": "Groom, Nichola L.", "username": "nichola.groom", "password": "vAesgS"},
    {"name": "Hall, Kalea", "username": "kalea.hall", "password": "4weehZ"},
    {"name": "Herbst, Svea A.", "username": "svea.herbst", "password": "rTV2O8"},
    {"name": "Hickman, Renee", "username": "renee.hickman", "password": "YqKDNk"},
    {"name": "Hood-Nuño, David", "username": "david.hood", "password": "ZWcmbx"},
    {"name": "Hu, Krystal", "username": "krystal.hu", "password": "YL4uog"},
    {"name": "Huffstutter, PJ", "username": "pj.huffstutter", "password": "m4I6Ea"},
    {"name": "Ingwersen, Julie R.", "username": "julie.ingwersen", "password": "vebx88"},
    {"name": "Jao, Nicole", "username": "nicole.jao", "password": "W6RcMP"},
    {"name": "Jeans, David", "username": "david.jeans", "password": "SBdIrS"},
    {"name": "Jones, Diana", "username": "diana.jones2", "password": "KgeMBZ"},
    {"name": "Kearney, Laila", "username": "laila.kearney", "password": "lerOfL"},
    {"name": "Kerber, Ross J.", "username": "ross.kerber", "password": "em8rZp"},
    {"name": "Khan, Shariq A.", "username": "shariq.khan", "password": "X8B7o1"},
    {"name": "Kirkham, Chris", "username": "chris.kirkham", "password": "0qxsfs"},
    {"name": "Knauth, Dietrich", "username": "dietrich.knauth", "password": "TrdWZ8"},
    {"name": "Koh, Gui Qing", "username": "guiqing.koh", "password": "8zE47J"},
    {"name": "Krauskopf, Lewis S.", "username": "lewis.krauskopf", "password": "W1KfYc"},
    {"name": "Landay, Jonathan S.", "username": "jonathan.landay", "password": "t3qK8O"},
    {"name": "Lang, Hannah", "username": "hannah.lang", "password": "6ZYhdE"},
    {"name": "Levine, Daniel R.", "username": "dan.levine", "password": "8gZVgu"},
    {"name": "Levy, Rachael", "username": "rachael.levy", "password": "5yxC6v"},
    {"name": "Lynch, Sarah N.", "username": "sarah.n.lynch", "password": "KGmRuz"},
    {"name": "MCLYMORE, ARRIANA", "username": "arriana.mclymore", "password": "P0Pr3p"},
    {"name": "Matthews, Laura", "username": "laura.matthews", "password": "uNDqpN"},
    {"name": "McCartney, Georgina", "username": "georgina.mccartney", "password": "1pOFnd"},
    {"name": "McCaskill, Nolan", "username": "nolan.mccaskill", "password": "HB3bZE"},
    {"name": "McGee, Suzanne", "username": "suzanne.mcgee", "password": "9tfPq1"},
    {"name": "McKay, Rich", "username": "rich.mckay", "password": "7SgXrm"},
    {"name": "McLaughlin, Timothy J.", "username": "tim.mclaughlin", "password": "zyvZHp"},
    {"name": "Mikolajczak, Chuck", "username": "charles.mikolajczak", "password": "Bxvsw4"},
    {"name": "Mutikani, Lucia V.", "username": "lucia.mutikani", "password": "6v9mKR"},
    {"name": "Nellis, Stephen", "username": "stephen.nellis", "password": "t5x582"},
    {"name": "Niasse, Amina", "username": "amina.niasse", "password": "Bk1pUJ"},
    {"name": "Oguh, Chibuike", "username": "chibuike.oguh", "password": "F25GdC"},
    {"name": "Oladipo, Doyinsola", "username": "doyinsola.oladipo", "password": "i3BfD0"},
    {"name": "Parraga, Marianna", "username": "marianna.parraga", "password": "6qOs8k"},
    {"name": "Paul, Katie", "username": "katie.paul", "password": "5quivt"},
    {"name": "Plume, Karl", "username": "karl.plume", "password": "cYjcuI"},
    {"name": "Polansek, Tom", "username": "thomas.polansek", "password": "YTZBXF"},
    {"name": "Prentice, Chris", "username": "christine.prentice", "password": "ThCpqP"},
    {"name": "Queen, Jack", "username": "jack.queen", "password": "h3FkGU"},
    {"name": "Randewich, Noel", "username": "noel.randewich", "password": "Oul6d4"},
    {"name": "Raymond, Nate", "username": "nate.raymond", "password": "QerXFR"},
    {"name": "Respaut, Robin", "username": "robin.respaut", "password": "kdD99a"},
    {"name": "Roulette, Joey", "username": "joey.roulette", "password": "IWvMoz"},
    {"name": "Roy, Abhirup", "username": "abhirup.roy", "password": "7PhxxA"},
    {"name": "Rozen, Courtney", "username": "courtney.rozen", "password": "qYbpog"},
    {"name": "Saphir, Ann", "username": "ann.saphir", "password": "L7K3GB"},
    {"name": "Scarcella, Mike", "username": "mike.scarcella", "password": "e1JTdM"},
    {"name": "Scheyder, Ernest", "username": "ernest.scheyder", "password": "zTLzxn"},
    {"name": "Schlitz, Heather", "username": "heather.schlitz", "password": "aZj4WW"},
    {"name": "Seba, Erwin", "username": "erwin.seba", "password": "V2YCbQ"},
    {"name": "Seetharaman, Deepa", "username": "deepa.seetharaman", "password": "PngRo2"},
    {"name": "Shepardson, David", "username": "david.shepardson", "password": "t0vKuh"},
    {"name": "Shirouzu, Norihiko", "username": "norihiko.shirouzu", "password": "OkfbMg"},
    {"name": "Singh, Rajesh Kumar", "username": "rajeshkumar.singh", "password": "zAihdK"},
    {"name": "Somasekhar, Arathy", "username": "arathy.s", "password": "WKlDQv"},
    {"name": "Spector, Mike", "username": "mike.spector", "password": "fiQUpm"},
    {"name": "Steenhuysen, Julie D.", "username": "julie.steenhuysen", "password": "WxgxB1"},
    {"name": "Stempel, Jonathan E.", "username": "jon.stempel", "password": "fBuzr8"},
    {"name": "Stone, Mike", "username": "mike.stone", "password": "9SVtNz"},
    {"name": "Summerville, Abigail", "username": "abigail.summerville", "password": "00XT4F"},
    {"name": "Teixeira, Marcelo", "username": "marcelo.teixeira", "password": "gMMg71"},
    {"name": "Terhune, Chad", "username": "chad.terhune", "password": "XOhG3D"},
    {"name": "Tracy, Matt", "username": "matt.tracy", "password": "rSteQt"},
    {"name": "Tsvetkova, Maria", "username": "maria.tsvetkova", "password": "JLhDPt"},
    {"name": "Valetkevitch, Caroline", "username": "caroline.valetkevitch", "password": "s3nXAc"},
    {"name": "Valle, Sabrina", "username": "sabrina.valle", "password": "SVALJL"},
    {"name": "Vicens, AJ", "username": "a.j.vicens", "password": "7ZmHvT"},
    {"name": "Vinn, Milana", "username": "milana.vinn", "password": "nXpjsd"},
    {"name": "Volcovici, Valerie", "username": "valerie.volcovici", "password": "Zd7nb9"},
    {"name": "Wang, Echo", "username": "e.wang", "password": "Z58njU"},
    {"name": "Wiessner, Daniel", "username": "daniel.wiessner", "password": "153gmC"},
    {"name": "Williams, Curtis", "username": "curtis.williams", "password": "K9yGJo"},
    {"name": "Wingrove, Patrick", "username": "patrick.wingrove", "password": "YZpmIK"},
    {"name": "Winter, Jana", "username": "jana.winter", "password": "ch6nBi"},
    {"name": "Wolfe, Jan", "username": "jan.wolfe", "password": "6aGKGY"},
    {"name": "Thomas, David", "username": "david.thomas", "password": "2nq5q6"},
    {"name": "Sloan, Karen", "username": "karen.sloan", "password": "s0lSlf"},
    {"name": "TEST ACCOUNT", "username": "test", "password": "test123"},
]

@app.route('/api/reload-reporters-from-csv', methods=['POST'])
def reload_reporters_from_csv():
    """Reload reporter accounts from embedded credentials data (ADMIN ONLY)"""
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        reporters = {}
        
        # Add admin account
        reporters['admin'] = {
            'name': 'Admin',
            'is_manager': True,
            'password': generate_password_hash('admin123')
        }
        
        # Add all reporters from embedded data
        for rep in REPORTER_CREDENTIALS:
            username = rep['username']
            name = rep['name']
            password = rep['password']
            
            reporters[username] = {
                'name': name,
                'is_manager': False,
                'password': generate_password_hash(password)
            }
        
        # Save to reporters.json
        save_json(REPORTERS_FILE, reporters)
        
        return jsonify({
            'success': True,
            'message': f'Successfully reloaded {len(reporters) - 1} reporter accounts',
            'total_accounts': len(reporters)
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reset-data', methods=['POST'])
def reset_data():
    """Reset preferences and assignments (ADMIN ONLY - for testing)"""
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        # Create backup before resetting
        create_auto_backup()
        
        # Clear preferences
        save_json(PREFERENCES_FILE, {})
        
        # Clear assignments
        save_json(ASSIGNMENTS_FILE, {})
        
        # Unlock preferences
        settings = get_settings()
        settings['is_locked'] = False
        save_json(SETTINGS_FILE, settings)
        
        return jsonify({
            'success': True,
            'message': 'All preferences and assignments cleared. System unlocked. Backup created.'
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/create-backup', methods=['POST'])
def trigger_backup():
    """Manually trigger a backup (ADMIN ONLY)"""
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    success = create_auto_backup()
    if success:
        return jsonify({
            'success': True,
            'message': 'Backup created successfully'
        })
    else:
        return jsonify({'error': 'Backup failed'}), 500

@app.route('/api/list-backups')
def list_backups():
    """List all available backups (ADMIN ONLY)"""
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        backup_files = sorted([f for f in os.listdir(BACKUP_DIR) if f.startswith('auto_backup_')], reverse=True)
        backups = []
        
        for filename in backup_files[:30]:  # Show last 30
            filepath = os.path.join(BACKUP_DIR, filename)
            stat = os.stat(filepath)
            backups.append({
                'filename': filename,
                'size': stat.st_size,
                'created': datetime.fromtimestamp(stat.st_mtime).isoformat()
            })
        
        return jsonify({
            'success': True,
            'backups': backups,
            'total': len(backup_files)
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/initialize-system', methods=['GET'])
def initialize_system():
    """PUBLIC ENDPOINT: Initialize reporters.json from embedded credentials (NO AUTH REQUIRED)"""
    try:
        reporters = {}
        
        # Add admin account
        reporters['admin'] = {
            'name': 'Admin',
            'is_manager': True,
            'password': generate_password_hash('admin123')
        }
        
        # Add all reporters from embedded data
        for rep in REPORTER_CREDENTIALS:
            username = rep['username']
            name = rep['name']
            password = rep['password']
            
            reporters[username] = {
                'name': name,
                'is_manager': False,
                'password': generate_password_hash(password)
            }
        
        # Save to reporters.json
        save_json(REPORTERS_FILE, reporters)
        
        return jsonify({
            'success': True,
            'message': f'Successfully initialized {len(reporters) - 1} reporter accounts + admin',
            'total_accounts': len(reporters),
            'note': 'You can now login with your credentials'
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)
